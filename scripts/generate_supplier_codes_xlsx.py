"""Generate FIAD_Supplier_Codes.xlsx from current production data.

Columns match the original file the user had:
  Booth · Company · Category · Login Passcode · Booth QR Token · Email · Contact · Facebook

Output goes to /Users/kylejarque/Documents/Claude/fiad/FIAD_Supplier_Codes.xlsx
(gitignored — share via private channel only).
"""
import json
import urllib.request
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

URL = "https://cjhnsyldnzdedgianzsj.supabase.co"
KEY = ("eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImNqaG5z"
       "eWxkbnpkZWRnaWFuenNqIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzkxNjg1NDIsImV4cCI6MjA5"
       "NDc0NDU0Mn0.A0BYsTrGpqVmXT6OdKndxytNuOoMJJmNGCYSTYrk48c")
OUT = "/Users/kylejarque/Documents/Claude/fiad/FIAD_Supplier_Codes.xlsx"

req = urllib.request.Request(
    f"{URL}/rest/v1/stores?select=name,booth_number,category,passcode,qr_token,email,contact,social_media&order=booth_number",
    headers={"apikey": KEY},
)
with urllib.request.urlopen(req) as r:
    stores = json.loads(r.read().decode())

print(f"Fetched {len(stores)} stores")

wb = Workbook()
ws = wb.active
ws.title = "Supplier Codes"

# ── Title row (merged across all columns) ──
ws.merge_cells("A1:H1")
ws["A1"] = "Forever in a Day — Supplier Login Codes & Booth QR Tokens"
ws["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", start_color="3E2A3E")  # plum
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# ── Header row ──
headers = ["Booth", "Company", "Category", "Login Passcode", "Booth QR Token",
           "Email", "Contact", "Facebook"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col, value=h)
    cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color="E63F75")  # coral
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 28

# ── Data rows ──
thin = Side(border_style="thin", color="EAEAEA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
alt_fill = PatternFill("solid", start_color="FFF7F4")  # very pale coral wash

for i, s in enumerate(stores, start=3):
    row = [
        s.get("booth_number") or "",
        s.get("name") or "",
        s.get("category") or "",
        s.get("passcode") or "",
        s.get("qr_token") or "",
        s.get("email") or "",
        s.get("contact") or "",
        s.get("social_media") or "",
    ]
    for col, val in enumerate(row, 1):
        cell = ws.cell(row=i, column=col, value=val)
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = border
        if i % 2 == 0:
            cell.fill = alt_fill
    # Bold the passcode for quick scanning
    ws.cell(row=i, column=4).font = Font(name="Arial", size=11, bold=True, color="3E2A3E")
    ws.row_dimensions[i].height = 24

# ── Column widths (tuned for readability without manual fiddling) ──
widths = [10, 32, 30, 14, 26, 32, 22, 36]
for col, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = w

# ── Freeze title + header rows so they stay visible while scrolling ──
ws.freeze_panes = "A3"

wb.save(OUT)
print(f"Wrote {OUT}")
print()
print("Summary by category:")
from collections import Counter
for cat, n in sorted(Counter(s.get("category") or "(none)" for s in stores).items()):
    print(f"  {n:3d}  {cat}")
